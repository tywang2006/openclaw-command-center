#!/usr/bin/env python3
"""Minimal document converter. Usage: convert.py <input> <output> <format>"""
import sys, json, csv, os

def main():
    if len(sys.argv) != 4:
        print(json.dumps({"error": "Usage: convert.py <input> <output> <format>"}))
        sys.exit(1)

    src, dst, fmt = sys.argv[1], sys.argv[2], sys.argv[3]
    ext = os.path.splitext(src)[1].lower()

    if ext == '.xlsx' and fmt == 'csv':
        import openpyxl
        wb = openpyxl.load_workbook(src)
        ws = wb.active
        with open(dst, 'w', newline='') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
    elif ext == '.xlsx' and fmt == 'json':
        import openpyxl
        wb = openpyxl.load_workbook(src)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, [v if v is not None else "" for v in row])))
        with open(dst, 'w') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    elif ext == '.docx' and fmt == 'txt':
        from docx import Document
        doc = Document(src)
        with open(dst, 'w') as f:
            f.write('\n'.join(p.text for p in doc.paragraphs))
    elif ext in ('.csv', '.txt', '.md', '.json') and fmt in ('txt', 'json', 'csv'):
        import shutil
        shutil.copy2(src, dst)
    else:
        print(json.dumps({"error": f"Conversion from {ext} to {fmt} not supported"}))
        sys.exit(1)

    print(json.dumps({"success": True}))

if __name__ == '__main__':
    main()
